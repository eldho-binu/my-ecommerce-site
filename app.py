from flask import Flask, render_template, request, session, redirect, url_for, flash
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key'

# Excel configuration
EXCEL_FILE = 'ecommerce_data.xlsx'

# Initialize Excel sheets if not exists
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()

        # Products Sheet
        ws1 = wb.active
        ws1.title = 'Products'
        ws1.append(["id", "name", "price", "image_url", "category"])

        # Users Sheet
        ws2 = wb.create_sheet('Users')
        ws2.append(["username", "password", "email", "phone"])

        # Orders Sheet
        ws3 = wb.create_sheet('Orders')
        ws3.append(["order_id", "username", "product_id", "product_name", "status"])

        wb.save(EXCEL_FILE)

# Read all products from Excel
def read_all_products():
    wb = load_workbook(EXCEL_FILE)
    sheet = wb['Products']
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            records.append({
                "id": row[0],
                "name": row[1],
                "price": row[2],
                "image_url": row[3],
                "category": row[4]
            })
    return records

# 🔹 Filter products by search query
def get_products(search=None):
    products = read_all_products()
    if search:
        return [p for p in products if search.lower() in p['name'].lower()]
    return products

# 🔹 Filter by category
@app.route('/category/<category_name>')
def get_products_by_category(category_name):
    products = read_all_products()
    filtered = [p for p in products if (p.get('category') or '').lower() == category_name.lower()]

    return render_template('category_list.html', products=filtered, category=category_name)

@app.route('/category')
def category():
    return render_template('category.html')

# 🔹 Home Page
@app.route('/')
def home():
    search = request.args.get('q')
    products = get_products(search)
    return render_template("index.html", products=products, search_done=bool(search))

# 🔹 Add to Cart
@app.route('/add-to-cart/<int:product_id>')
def add_to_cart(product_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    cart = session.get('cart', [])
    if product_id not in cart:
        cart.append(product_id)
    session['cart'] = cart
    return redirect(url_for('home'))

# 🔹 View Cart
@app.route('/cart')
def cart():
    if 'username' not in session:
        return redirect(url_for('login'))

    cart_ids = session.get('cart', [])
    all_products = read_all_products()
    products = [p for p in all_products if p['id'] in cart_ids]
    return render_template("cart.html", products=products)

def read_order():
    wb = load_workbook(EXCEL_FILE)
    sheet = wb['Orders']
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            records.append({
                "order_id": row[0],
                "username": row[1],
                "product_id": row[2],
                "product_name": row[3],
                "status": row[4]
            })
    return records

@app.route('/my-orders')
def orders():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    products = read_order()
    filtered = [p for p in products if (p.get('username') or '').lower() == username.lower()]

    return render_template('my_orders.html', products=filtered, username=username)


# 🔹 Login using Excel
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        wb = load_workbook(EXCEL_FILE)
        sheet = wb['Users']

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                session['username'] = username
                flash('Login successful!', 'success')
                return redirect(url_for('home'))

        flash('Invalid username or password', 'danger')

    return render_template('login.html')

# 🔹 Signup
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        phone = request.form['phone']

        wb = load_workbook(EXCEL_FILE)
        sheet = wb['Users']

        # Check if user exists
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                flash('Username already exists', 'warning')
                return redirect(url_for('signup'))

        sheet.append([username, password, email, phone])
        wb.save(EXCEL_FILE)

        flash('Account created successfully. Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('signup.html')

# 🔹 Track Order
@app.route('/track', methods=['GET', 'POST'])
def track_order():
    if request.method == 'POST':
        order_id = request.form['order_id']
        wb = load_workbook(EXCEL_FILE)
        sheet = wb['Orders']

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == order_id:
                result = {
                    'order_id': row[0],
                    'user_name': row[1],
                    'product_name': row[3],
                    'status': row[4]
                }
                return render_template('track.html', result=result)

        return render_template('track.html', error="Order ID not found.")

    return render_template('track.html')

# 🔹 Dummy Checkout
@app.route('/checkout')
def checkout():
    return "✅ This is a demo checkout page."

# 🔹 Logout
@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out successfully!", "info")
    return redirect(url_for('home'))

# 🔹 Initialize Excel and Run App
if __name__ == '__main__':
    initialize_excel()
    app.run(debug=True)
